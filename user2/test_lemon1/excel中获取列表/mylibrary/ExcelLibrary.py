from openpyxl import load_workbook
from robot.api import logger

"""
1.本扩展库只适合excel模板按照顺序为
case_id	case_name	function_name	url	data	expected	actual	result
2.只支持openpyxl支持的文件格式
"""


class ExcelLibrary(object):
    ROBOT_LIBRARY_VERSION = 1.0
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    """
	Test library for testing *Calculator* business logic.
    Interacts with the calculator directly using its ``push`` method.
    """

    def __init__(self):
		
        logger.info("doexcel init")
        pass


    def get_cases(self,filename, sheetname):

        self.workbook = load_workbook(filename)
        self.sheet =self.workbook[sheetname]
        max_row =self. sheet.max_row
        cases = []
        for r in range(2, max_row + 1):
            case = {}
            case['case_id'] = self.sheet.cell(r, column=1).value
            case['title'] = self.sheet.cell(r, column=2).value
            case['method'] = self.sheet.cell(row=r, column=3).value
            case['url'] = self.sheet.cell(row=r, column=4).value
            case['data'] = self.sheet.cell(row=r, column=5).value
            case['expected'] = self.sheet.cell(row=r, column=6).value
            case['sql'] = self.sheet.cell(row=r, column=9).value
            cases.append(case)
		
        self.workbook.close()
        return cases


    def writeResult(self,filename,sheetname, row, actual, result):
		
        self.workbook = load_workbook(filename)
        sheet = self.workbook[sheetname]
        sheet.cell(int(row), 7).value = actual
        sheet.cell(int(row), 8).value = result
        self.workbook.save(filename=filename)
        self.workbook.close()


if __name__ == '__main__':
    do = ExcelLibrary()
    s=do.get_cases('D:\Api_liliang\case1.xlsx','verifyUserAuth')
    print(s)